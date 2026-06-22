from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from alunos.models import GrupoMusical, Aluno
from professores.models import EnsaiosproModel
from .models import Partitura
from .forms import PartituraForm, GrupoForm


def editar_funcao(request, aluno_id):
    aluno = get_object_or_404(Aluno, id=aluno_id)

    if request.method == "POST":
        aluno.funcao = request.POST.get("funcao")
        aluno.save()
        return redirect("grupos:detalhe_grupo", id=aluno.grupo.id)

    return render(request, "grupos/editar_funcao.html", {"aluno": aluno})


@login_required
def grupo_detalhes(request, id):
    grupo = get_object_or_404(GrupoMusical, id=id)
    musicos = Aluno.objects.filter(grupo=grupo)
    partituras = grupo.partituras.all()

    contexto = {
        "grupo": grupo,
        "musicos": musicos,
        "partituras": partituras
    }

    # Aluno vê template com sidebar de aluno; professor/gerente vêem o template normal
    tipo = getattr(getattr(request.user, 'perfil', None), 'tipo', None)
    if tipo == 'aluno':
        return render(request, "grupos/detalhes_aluno.html", contexto)
    return render(request, "grupos/detalhes.html", contexto)

@login_required
def enviar_partitura(request, id):
    grupo = get_object_or_404(GrupoMusical, id=id)

    if request.method == "POST":
        form = PartituraForm(request.POST, request.FILES)
        if form.is_valid():
            partitura = form.save(commit=False)
            partitura.grupo = grupo
            partitura.save()
            return redirect("grupos:detalhes", id=grupo.id)

    else:
        form = PartituraForm()

    return render(request, "grupos/enviar_partitura.html", {
        "form": form,
        "grupo": grupo
    })



@login_required
def lista_grupos(request):
    grupos = GrupoMusical.objects.all()
    return render(request, "grupos/lista.html", {"grupos": grupos})

@login_required
def grupos_listar(request):
    grupos = GrupoMusical.objects.all()
    ativos_count = grupos.filter(ativo=True).count()
    membros_totais_count = Aluno.objects.filter(grupo__isnull=False).count()
    
    hoje = timezone.now().date()
    proximos_ensaios_count = EnsaiosproModel.objects.filter(data__gte=hoje, cancelado=False).count()
    
    contexto = {
        "grupos": grupos,
        "ativos_count": ativos_count,
        "membros_totais_count": membros_totais_count,
        "proximos_ensaios_count": proximos_ensaios_count,
    }
    return render(request, "grupos/listar.html", contexto)


@login_required
def grupo_criar(request):
    if request.method == "POST":
        form = GrupoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("grupos:listar")
    else:
        form = GrupoForm()

    return render(request, "grupos/criar.html", {"form": form})


@login_required
def grupo_editar(request, id):
    grupo = get_object_or_404(GrupoMusical, id=id)
    
    if request.method == "POST":
        form = GrupoForm(request.POST, instance=grupo)
        if form.is_valid():
            form.save()
            return redirect("grupos:listar")
    else:
        form = GrupoForm(instance=grupo)

    return render(request, "grupos/editar.html", {"form": form, "grupo": grupo})


@login_required
def grupo_arquivar(request, id):
    grupo = get_object_or_404(GrupoMusical, id=id)
    grupo.ativo = False
    grupo.save()
    return redirect("grupos:listar")


@login_required
def grupo_reativar(request, id):
    grupo = get_object_or_404(GrupoMusical, id=id)
    grupo.ativo = True
    grupo.save()
    return redirect("grupos:listar")


import csv
from django.http import HttpResponse

@login_required
def exportar_grupos_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="grupos.csv"'
    
    response.write(b'\xEF\xBB\xBF')
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Nome', 'Descricao', 'Ativo', 'Integrantes'])
    
    grupos = GrupoMusical.objects.all()
    for g in grupos:
        ativo = 'Sim' if g.ativo else 'Não'
        integrantes = Aluno.objects.filter(grupo=g)
        integrantes_str = ", ".join([f"{a.nome} {a.sobrenome}" for a in integrantes])
        writer.writerow([
            g.nome,
            g.descricao,
            ativo,
            integrantes_str
        ])
    return response


@login_required
def exportar_grupos_excel(request):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Grupos Musicais"
    
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="8B6E2F")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B6E2F", end_color="8B6E2F", fill_type="solid")
    data_font = Font(name=font_family, size=11)
    
    thin_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    ws['A1'] = "Relatório de Grupos Musicais — ABANFAR"
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 30
    
    headers = ['Nome do Grupo', 'Descrição', 'Ativo?', 'Integrantes']
    ws.append([])
    ws.append(headers)
    
    ws.row_dimensions[3].height = 25
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    grupos = GrupoMusical.objects.all()
    group_sizes = []
    
    for g in grupos:
        ativo = 'Sim' if g.ativo else 'Não'
        integrantes = Aluno.objects.filter(grupo=g)
        integrantes_str = ", ".join([f"{a.nome} {a.sobrenome}" for a in integrantes])
        
        row_data = [
            g.nome,
            g.descricao if g.descricao else 'Sem descrição',
            ativo,
            integrantes_str
        ]
        ws.append(row_data)
        group_sizes.append((g.nome, len(integrantes)))
        
    last_row = len(grupos) + 3
    for r in range(4, last_row + 1):
        ws.row_dimensions[r].height = 20
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.border = thin_border
            if c == 3:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    ws['F2'] = "Grupo"
    ws['G2'] = "Integrantes"
    ws['F2'].font = header_font
    ws['F2'].fill = header_fill
    ws['F2'].alignment = center_align
    ws['G2'].font = header_font
    ws['G2'].fill = header_fill
    ws['G2'].alignment = center_align
    
    sum_row = 3
    for g_name, size in group_sizes:
        ws.cell(row=sum_row, column=6, value=g_name).font = data_font
        ws.cell(row=sum_row, column=6).border = thin_border
        ws.cell(row=sum_row, column=7, value=size).font = data_font
        ws.cell(row=sum_row, column=7).alignment = center_align
        ws.cell(row=sum_row, column=7).border = thin_border
        sum_row += 1
        
    if len(group_sizes) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Integrantes por Grupo Musical"
        chart.y_axis.title = "Membros"
        chart.x_axis.title = "Grupo"
        
        data_ref = Reference(ws, min_col=7, min_row=2, max_row=sum_row-1)
        cats_ref = Reference(ws, min_col=6, min_row=3, max_row=sum_row-1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.legend = None
        chart.width = 16
        chart.height = 11
        ws.add_chart(chart, "I2")
        
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].column <= 4:
            max_len = 0
            for cell in col:
                if cell.row == 1 and col_letter == 'A':
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            if col_letter == 'D':
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
            else:
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="grupos.xlsx"'
    return response


@login_required
def exportar_grupos_pdf(request):
    from contas.utils import render_to_pdf
    from django.utils import timezone
    
    grupos = GrupoMusical.objects.all()
    total_grupos = grupos.count()
    ativos = GrupoMusical.objects.filter(ativo=True).count()
    inativos = total_grupos - ativos
    
    grupos_with_members = []
    for g in grupos:
        members = Aluno.objects.filter(grupo=g)
        grupos_with_members.append({
            'obj': g,
            'members': members,
            'members_count': members.count()
        })
        
    context = {
        'grupos': grupos_with_members,
        'total_grupos': total_grupos,
        'ativos': ativos,
        'inativos': inativos,
        'data_geracao': timezone.now().strftime("%d/%m/%Y %H:%M"),
        'titulo': 'Relatório Geral de Grupos Musicais'
    }
    
    pdf_response = render_to_pdf('grupos/grupos_pdf.html', context)
    if pdf_response:
        pdf_response['Content-Disposition'] = 'attachment; filename="grupos.pdf"'
        return pdf_response
    return HttpResponse("Erro ao gerar PDF", status=500)

