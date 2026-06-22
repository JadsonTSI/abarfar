from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.models import User
from .models import Perfil
from alunos.models import Aluno, GrupoMusical, Naipe
from professores.models import ProfessorModel, EnsaiosproModel, ApresentacaoModel
from materias.models import Materia, MatriculaMateria
from instrumentos.models import Instrumento, InstrumentoEmprestimo, IoTScan
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .decorators import gerente_required, professor_required, aluno_required
from django.utils import timezone
import datetime
import json


def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        senha = request.POST.get("senha")

        user = authenticate(request, username=username, password=senha)

        if user:
            login(request, user)
            perfil = Perfil.objects.get(user=user)

            if perfil.tipo == "gerente":
                return redirect("contas:painel_gerente")
            elif perfil.tipo == "professor":
                return redirect("professores:professoreshome")
            elif perfil.tipo == "aluno":
                return redirect("contas:painel_aluno")

        messages.error(request, "Usuário ou senha incorretos!")
        return redirect("contas:login")

    return render(request, "contas/login.html")


def logout_view(request):
    logout(request)
    return redirect("contas:login")


@gerente_required
def painel_gerente(request):
    from django.db.models import Q
    
    # 0. Busca Global
    busca = request.GET.get("busca", "").strip()
    resultados_busca = None
    if busca:
        resultados_busca = {
            "alunos": Aluno.objects.filter(
                Q(nome__icontains=busca) |
                Q(sobrenome__icontains=busca) |
                Q(matricula__icontains=busca)
            ).select_related('instrumento', 'grupo'),
            "professores": ProfessorModel.objects.filter(
                Q(nome__icontains=busca) |
                Q(instrumento__icontains=busca)
            ),
            "instrumentos": Instrumento.objects.filter(
                Q(nome__icontains=busca) |
                Q(identificador__icontains=busca)
            ),
            "grupos": GrupoMusical.objects.filter(
                Q(nome__icontains=busca)
            )
        }

    # 1. Estatísticas dos Cards
    total_alunos = Aluno.objects.count()
    total_professores = ProfessorModel.objects.count()
    total_instrumentos = Instrumento.objects.count()
    total_grupos = GrupoMusical.objects.count()
    
    # 2. Alunos Recentes (últimos 5)
    alunos_recentes = Aluno.objects.all().select_related('instrumento', 'grupo').order_by('-id')[:5]
    
    # 3. Estatísticas dos Naipes para os Progress Bars (Projetos / Naipes)
    naipes = Naipe.objects.all()
    naipes_data = []
    max_count = 0
    
    for n in naipes:
        cnt = Aluno.objects.filter(naipe=n).count()
        naipes_data.append({"nome": n.nome, "quantidade": cnt})
        if cnt > max_count:
            max_count = cnt
            
    # Calcular porcentagem
    for nd in naipes_data:
        nd["percent"] = int((nd["quantidade"] / max_count * 100)) if max_count > 0 else 0
        
    # Ordenar por quantidade de alunos decrescente
    naipes_data = sorted(naipes_data, key=lambda x: x["quantidade"], reverse=True)[:4]
    
    # 4. Próximos Eventos (Ensaios futuros)
    hoje = timezone.now().date()
    eventos = EnsaiosproModel.objects.filter(data__gte=hoje, cancelado=False).order_by('data', 'inicio')[:3]
    
    # Formatar eventos para o template
    eventos_formatados = []
    meses = {
        1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
        7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez"
    }
    for ev in eventos:
        eventos_formatados.append({
            "dia": ev.data.strftime("%d"),
            "mes": meses.get(ev.data.month, "Mês"),
            "titulo": ev.nome,
            "local": ev.local,
            "hora": ev.inicio.strftime("%H:%M")
        })
        
    # 5. Notificações / Alertas recentes
    alertas = []
    prestamos_ativos = InstrumentoEmprestimo.objects.filter(devolvido=False).select_related('instrumento', 'aluno')
    for p in prestamos_ativos:
        dias = (hoje - p.data_emprestimo).days
        if dias > 14:
            alertas.append({
                "id": f"v-{p.id}",
                "grave": True,
                "msg": f"Prazo vencido: {p.instrumento.nome} — {p.aluno.nome} (+{dias - 14} dias)",
                "tempo": f"Há {dias} dias"
            })
            
    devolvidos_recentes = InstrumentoEmprestimo.objects.filter(devolvido=True).select_related('instrumento', 'aluno').order_by('-data_devolucao')[:3]
    for r in devolvidos_recentes:
        alertas.append({
            "id": f"d-{r.id}",
            "grave": False,
            "msg": f"Instrumento {r.instrumento.nome} retornado pelo aluno {r.aluno.nome}.",
            "tempo": "Recentemente"
        })
        
    alertas = alertas[:4]
    
    # 6. Dados para os gráficos
    total_inst_ativos = Instrumento.objects.filter(ativo=True).count()
    com_rfid = Instrumento.objects.filter(ativo=True, rfid__isnull=False).exclude(rfid="").count()
    sem_rfid = total_inst_ativos - com_rfid
    
    emprestados = InstrumentoEmprestimo.objects.filter(devolvido=False).count()
    disponiveis = total_inst_ativos - emprestados
    
    # 6.1. Condições dos Instrumentos
    cond_otimo = Instrumento.objects.filter(condicao='otimo').count()
    cond_bom = Instrumento.objects.filter(condicao='bom').count()
    cond_regular = Instrumento.objects.filter(condicao='regular').count()
    cond_ruim = Instrumento.objects.filter(condicao='ruim').count()
    
    # 6.2. Alunos por Grupo Musical
    grupos_labels = []
    grupos_counts = []
    for g in GrupoMusical.objects.all():
        cnt = Aluno.objects.filter(grupo=g).count()
        if cnt > 0:
            grupos_labels.append(g.nome)
            grupos_counts.append(cnt)
            
    # 6.3. Histórico de Empréstimos vs Scans IoT (Últimos 7 dias)
    labels_atividade = []
    emprestimos_counts = []
    scans_counts = []
    for i in range(6, -1, -1):
        dia = hoje - datetime.timedelta(days=i)
        labels_atividade.append(dia.strftime("%d/%b"))
        
        cnt_emp = InstrumentoEmprestimo.objects.filter(data_emprestimo=dia).count()
        emprestimos_counts.append(cnt_emp)
        
        cnt_scan = IoTScan.objects.filter(timestamp__date=dia).count()
        scans_counts.append(cnt_scan)
        
    context = {
        "total_alunos": total_alunos,
        "total_professores": total_professores,
        "total_instrumentos": total_instrumentos,
        "total_grupos": total_grupos,
        "alunos_recentes": alunos_recentes,
        "naipes_data": naipes_data,
        "eventos": eventos_formatados,
        "alertas": alertas,
        
        "chart_instrumentos": {
            "disponiveis": disponiveis,
            "emprestados": emprestados,
        },
        "chart_rfid": {
            "com_rfid": com_rfid,
            "sem_rfid": sem_rfid,
        },
        "chart_condicoes": {
            "labels": ["Ótimo", "Bom", "Regular", "Ruim"],
            "data": [cond_otimo, cond_bom, cond_regular, cond_ruim]
        },
        "chart_grupos": {
            "labels": grupos_labels,
            "data": grupos_counts
        },
        "chart_atividade": {
            "labels": labels_atividade,
            "emprestimos": emprestimos_counts,
            "scans": scans_counts
        },
        "busca": busca,
        "resultados_busca": resultados_busca,
    }
    return render(request, "contas/gerente_home.html", context)


@professor_required
def painel_professor(request):
    return redirect("professores:professoreshome")


@aluno_required
def painel_aluno(request):
    perfil = request.user.perfil
    aluno = perfil.aluno
    
    total_materias = MatriculaMateria.objects.filter(aluno=aluno).count()
    total_ensaios = EnsaiosproModel.objects.filter(cancelado=False).count()
    total_apresentacoes = ApresentacaoModel.objects.count()
    total_colegas = Aluno.objects.filter(naipe=aluno.naipe).exclude(id=aluno.id).count() if aluno.naipe else 0
    
    # Gráficos criativos / reais do Aluno
    chart_habilidades = {
        "labels": ["Ritmo", "Afinação", "Teoria", "Prática", "Presença"],
        "data": [85, 90, min(60 + (total_materias * 12), 100), 80, 95]
    }
    
    chart_jornada = {
        "labels": ["Matérias", "Próximos Ensaios", "Colegas de Naipe"],
        "data": [total_materias, total_ensaios, total_colegas]
    }
    
    context = {
        "perfil": perfil,
        "aluno": aluno,
        "total_materias": total_materias,
        "total_ensaios": total_ensaios,
        "total_apresentacoes": total_apresentacoes,
        "total_colegas": total_colegas,
        "chart_habilidades": chart_habilidades,
        "chart_jornada": chart_jornada,
    }
    return render(request, "contas/aluno_home.html", context)



@gerente_required
def lista_alunos(request):
    alunos = Aluno.objects.all().select_related('perfil__user', 'instrumento')
    
    ativos_count = alunos.filter(perfil__user__is_active=True).count()
    com_instrumento_count = alunos.filter(instrumento__isnull=False).count()
    
    # Novos este ano letivo (Julho a Julho)
    hoje = timezone.now().date()
    if hoje.month >= 7:
        ano_inicio = hoje.year
    else:
        ano_inicio = hoje.year - 1
        
    import datetime
    data_inicio_ciclo = datetime.date(ano_inicio, 7, 1)
    novos_este_mes_count = alunos.filter(perfil__user__date_joined__date__gte=data_inicio_ciclo).count()
    
    contexto = {
        "alunos": alunos,
        "ativos_count": ativos_count,
        "com_instrumento_count": com_instrumento_count,
        "novos_este_mes_count": novos_este_mes_count,
    }
    return render(request, "contas/gerente_lista_alunos.html", contexto)



@login_required
def perfil_view(request):
    perfil = request.user.perfil

    aluno = None
    professor = None
    materias = []

    # --- Se for aluno ---
    try:
        aluno = perfil.aluno
        materias = Materia.objects.filter(
            id__in=MatriculaMateria.objects.filter(aluno=aluno).values("materia")
        )
    except:
        aluno = None

    # --- Se for professor ---
    try:
        professor = perfil.professor  # ✅ AGORA ESTÁ CORRETO
        materias = Materia.objects.filter(professor=professor)
    except:
        professor = None

    contexto = {
        "perfil": perfil,
        "aluno": aluno,
        "professor": professor,
        "materias": materias,
    }

    return render(request, "contas/perfil.html", contexto)


@login_required
def alterar_foto(request):
    perfil = request.user.perfil

    if request.method == "POST" and request.FILES.get("foto"):
        foto = request.FILES["foto"]
        limit_mb = 5
        if foto.size > limit_mb * 1024 * 1024:
            messages.error(request, f"A imagem é muito grande. O limite máximo permitido é de {limit_mb}MB.")
        else:
            perfil.foto = foto
            perfil.save()
            messages.success(request, "Foto de perfil atualizada com sucesso.")

    return redirect("contas:perfil")

@csrf_exempt
def login_api(request):
    if request.method == "POST":
        try:
            data = request.POST
            username = data.get("username")
            senha = data.get("senha")
        except:
            return JsonResponse({"erro": "Dados invalidos"}, status=400)

        user = authenticate(request, username=username, password=senha)

        if user:
            login(request, user)
            perfil = Perfil.objects.get(user=user)
            return JsonResponse({
                "sucesso": True,
                "tipo": perfil.tipo,
                "username": user.username,
            })

        return JsonResponse({"erro": "Usuario ou senha incorretos"}, status=401)

    return JsonResponse({"erro": "Metodo nao permitido"}, status=405)


import csv
from django.http import HttpResponse

@gerente_required
def exportar_alunos_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="alunos.csv"'
    
    # Write BOM for Excel compatibility with UTF-8
    response.write(b'\xEF\xBB\xBF')
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Nome', 'Sobrenome', 'Matricula', 'Telefone', 'Instrumento', 'Naipe', 'Grupo', 'Status'])
    
    alunos = Aluno.objects.all().select_related('perfil__user', 'instrumento', 'naipe', 'grupo')
    for a in alunos:
        status = 'Ativo' if a.perfil.user.is_active else 'Inativo'
        writer.writerow([
            a.nome,
            a.sobrenome,
            a.matricula,
            a.telefone,
            a.instrumento.nome if a.instrumento else 'Sem instrumento',
            a.naipe.nome if a.naipe else 'Sem naipe',
            a.grupo.nome if a.grupo else 'Sem grupo',
            status
        ])
    return response


@gerente_required
def exportar_alunos_excel(request):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, Reference
    from django.db.models import Count
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Alunos"
    
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="8B6E2F")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B6E2F", end_color="8B6E2F", fill_type="solid")
    data_font = Font(name=font_family, size=11)
    
    thin_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    ws['A1'] = "Relatório de Alunos — ABANFAR"
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 30
    
    headers = ['Nome', 'Sobrenome', 'Matrícula', 'Telefone', 'Instrumento', 'Naipe', 'Grupo', 'Status']
    ws.append([])
    ws.append(headers)
    
    ws.row_dimensions[3].height = 25
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    alunos = Aluno.objects.all().select_related('perfil__user', 'instrumento', 'naipe', 'grupo')
    
    for a in alunos:
        status = 'Ativo' if a.perfil.user.is_active else 'Inativo'
        row_data = [
            a.nome,
            a.sobrenome,
            a.matricula,
            a.telefone,
            a.instrumento.nome if a.instrumento else 'Sem instrumento',
            a.naipe.nome if a.naipe else 'Sem naipe',
            a.grupo.nome if a.grupo else 'Sem grupo',
            status
        ]
        ws.append(row_data)
        
    last_row = len(alunos) + 3
    for r in range(4, last_row + 1):
        ws.row_dimensions[r].height = 20
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.border = thin_border
            if c in [3, 4, 8]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    ws['K2'] = "Naipe"
    ws['L2'] = "Quantidade"
    ws['K2'].font = header_font
    ws['K2'].fill = header_fill
    ws['K2'].alignment = center_align
    ws['L2'].font = header_font
    ws['L2'].fill = header_fill
    ws['L2'].alignment = center_align
    
    naipe_counts = Aluno.objects.values('naipe__nome').annotate(count=Count('id')).order_by('-count')
    sum_row = 3
    for nc in naipe_counts:
        naipe_name = nc['naipe__nome'] if nc['naipe__nome'] else 'Sem naipe'
        ws.cell(row=sum_row, column=11, value=naipe_name).font = data_font
        ws.cell(row=sum_row, column=11).border = thin_border
        ws.cell(row=sum_row, column=12, value=nc['count']).font = data_font
        ws.cell(row=sum_row, column=12).alignment = center_align
        ws.cell(row=sum_row, column=12).border = thin_border
        sum_row += 1
        
    if len(naipe_counts) > 0:
        pie = PieChart()
        labels = Reference(ws, min_col=11, min_row=3, max_row=sum_row-1)
        data = Reference(ws, min_col=12, min_row=2, max_row=sum_row-1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Distribuição de Alunos por Naipe"
        pie.width = 16
        pie.height = 11
        ws.add_chart(pie, "N2")
        
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].column <= 8:
            max_len = 0
            for cell in col:
                if cell.row == 1 and col_letter == 'A':
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="alunos.xlsx"'
    return response


@gerente_required
def exportar_alunos_pdf(request):
    from .utils import render_to_pdf
    
    alunos = Aluno.objects.all().select_related('perfil__user', 'instrumento', 'naipe', 'grupo')
    total_alunos = alunos.count()
    ativos = Aluno.objects.filter(perfil__user__is_active=True).count()
    inativos = total_alunos - ativos
    
    context = {
        'alunos': alunos,
        'total_alunos': total_alunos,
        'ativos': ativos,
        'inativos': inativos,
        'data_geracao': timezone.now().strftime("%d/%m/%Y %H:%M"),
        'titulo': 'Relatório Geral de Alunos'
    }
    
    pdf_response = render_to_pdf('contas/alunos_pdf.html', context)
    if pdf_response:
        pdf_response['Content-Disposition'] = 'attachment; filename="alunos.pdf"'
        return pdf_response
    return HttpResponse("Erro ao gerar PDF", status=500)