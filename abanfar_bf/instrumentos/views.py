from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from contas.decorators import gerente_required, iot_api_required
from .models import Instrumento, InstrumentoEmprestimo, IoTScan
from .forms import InstrumentoForm, InstrumentoEmprestimoForm

# LISTAR INSTRUMENTOS
@login_required
def list_instrumentos(request):
    instrumentos = Instrumento.objects.all()
    ativos_count = instrumentos.filter(ativo=True).count()
    da_associacao_count = instrumentos.filter(pertence_associacao=True).count()
    dos_alunos_count = instrumentos.filter(pertence_associacao=False).count()
    
    contexto = {
        "instrumentos": instrumentos,
        "ativos_count": ativos_count,
        "da_associacao_count": da_associacao_count,
        "dos_alunos_count": dos_alunos_count,
    }
    return render(request, "instrumentos/list.html", contexto)

# CADASTRAR INSTRUMENTO
@gerente_required
def criar_instrumento(request):
    if request.method == "POST":
        form = InstrumentoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("instrumentos:list")
    else:
        form = InstrumentoForm()

    return render(request, "instrumentos/form.html", {"form": form})

# ATRIBUIR / EMPRESTAR INSTRUMENTO
@gerente_required
def atribuir_instrumento(request, id):
    instrumento = get_object_or_404(Instrumento, id=id)

    if request.method == "POST":
        form = InstrumentoEmprestimoForm(request.POST)
        if form.is_valid():
            emprestimo = form.save(commit=False)
            emprestimo.instrumento = instrumento
            emprestimo.save()
            return redirect("instrumentos:list")
    else:
        form = InstrumentoEmprestimoForm()

    return render(request, "instrumentos/atribuir.html", {
        "form": form,
        "instrumento": instrumento
    })

# USO / HISTÓRICO DE EMPRESTIMOS
@login_required
def uso_instrumento(request, id):
    instrumento = get_object_or_404(Instrumento, id=id)
    usos = InstrumentoEmprestimo.objects.filter(instrumento=instrumento)

    return render(request, "instrumentos/uso.html", {
        "instrumento": instrumento,
        "usos": usos
    })


@gerente_required
def editar_instrumento(request, id):
    instrumento = get_object_or_404(Instrumento, id=id)
    
    if request.method == "POST":
        form = InstrumentoForm(request.POST, instance=instrumento)
        if form.is_valid():
            form.save()
            return redirect("instrumentos:list")
    else:
        form = InstrumentoForm(instance=instrumento)

    return render(request, "instrumentos/editar.html", {"form": form, "instrumento": instrumento})


@gerente_required
def excluir_instrumento(request, id):
    instrumento = get_object_or_404(Instrumento, id=id)

    if request.method == "POST":
        instrumento.delete()
        return redirect("instrumentos:list")

    return render(request, "instrumentos/excluir.html", {"instrumento": instrumento})


# ── ENDPOINTS DE API PARA O APLICATIVO ─────────────────────────────────────────

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
import json

@csrf_exempt
@iot_api_required
def api_painel_geral(request):
    # Total de instrumentos
    total_inst = Instrumento.objects.count()
    ativos_inst = Instrumento.objects.filter(ativo=True).count()
    inativos_inst = Instrumento.objects.filter(ativo=False).count()
    
    # Contagem de empréstimos
    emprestados_ids = InstrumentoEmprestimo.objects.filter(devolvido=False).values_list('instrumento_id', flat=True)
    emprestados_count = len(emprestados_ids)
    disponiveis_count = Instrumento.objects.filter(ativo=True).exclude(id__in=emprestados_ids).count()
    
    hoje = timezone.now().date()
    vencidos_count = 0
    prestamos_ativos = InstrumentoEmprestimo.objects.filter(devolvido=False)
    for p in prestamos_ativos:
        dias = (hoje - p.data_emprestimo).days
        if dias > 14:
            vencidos_count += 1
            
    ativos_prestamos = prestamos_ativos.count() - vencidos_count
    devolvidos_prestamos = InstrumentoEmprestimo.objects.filter(devolvido=True).count()
    total_prestamos = InstrumentoEmprestimo.objects.count()
    
    # Estatísticas de Alunos
    from alunos.models import Aluno, Naipe
    total_alunos = Aluno.objects.count()
    naipes = Naipe.objects.all()
    naipes_dict = {}
    for n in naipes:
        naipes_dict[n.nome] = Aluno.objects.filter(naipe=n).count()
        
    # Estatísticas de Ensaios
    from professores.models import EnsaiosproModel
    total_ensaios = EnsaiosproModel.objects.count()
    ativos_ensaios = EnsaiosproModel.objects.filter(cancelado=False).count()
    cancelados_ensaios = EnsaiosproModel.objects.filter(cancelado=True).count()
    
    # Lista de alertas
    alertas = []
    # Alertas de vencidos
    for p in prestamos_ativos:
        dias = (hoje - p.data_emprestimo).days
        if dias > 14:
            alertas.append({
                "id": f"v-{p.id}",
                "grave": True,
                "msg": f"Prazo vencido: {p.instrumento.nome} — {p.aluno.nome} (+{dias - 14} dias)"
            })
    # Devoluções recentes
    devolvidos_recentes = InstrumentoEmprestimo.objects.filter(devolvido=True).order_by('-data_devolucao')[:3]
    for r in devolvidos_recentes:
        alertas.append({
            "id": f"d-{r.id}",
            "grave": False,
            "msg": f"{r.instrumento.nome} devolvido por {r.aluno.nome}"
        })
    # Próximos ensaios hoje/amanhã
    ensaios_hoje = EnsaiosproModel.objects.filter(data__gte=hoje, cancelado=False).order_by('data', 'inicio')[:3]
    for e in ensaios_hoje:
        dia_str = "Hoje" if e.data == hoje else e.data.strftime("%d/%m/%Y")
        alertas.append({
            "id": f"e-{e.id}",
            "grave": False,
            "msg": f"Ensaio {e.nome} {dia_str} às {e.inicio.strftime('%H:%M')} — {e.local}"
        })
        
    # Próximo ensaio em destaque
    prox_e = EnsaiosproModel.objects.filter(data__gte=hoje, cancelado=False).order_by('data', 'inicio').first()
    prox_ensaio_data = None
    if prox_e:
        dia_str = "Hoje" if prox_e.data == hoje else prox_e.data.strftime("%d/%m/%Y")
        prox_ensaio_data = {
            "nome": prox_e.nome,
            "data": prox_e.data.strftime("%d/%m/%Y"),
            "inicio": prox_e.inicio.strftime("%H:%M"),
            "local": prox_e.local,
            "dia": dia_str
        }

    # Condições dos Instrumentos
    cond_otimo = Instrumento.objects.filter(condicao='otimo').count()
    cond_bom = Instrumento.objects.filter(condicao='bom').count()
    cond_regular = Instrumento.objects.filter(condicao='regular').count()
    cond_ruim = Instrumento.objects.filter(condicao='ruim').count()

    # Histórico de Empréstimos vs Scans IoT (Últimos 7 dias)
    import datetime
    labels_atividade = []
    emprestimos_counts = []
    scans_counts = []
    for i in range(6, -1, -1):
        dia = hoje - datetime.timedelta(days=i)
        labels_atividade.append(dia.strftime("%d/%b"))
        cnt_emp = InstrumentoEmprestimo.objects.filter(data_emprestimo=dia).count()
        emprestimos_counts.append(cnt_emp)
        cnt_scan = IoTScan.objects.filter(timestamp__date=dia).count()
        scans_counts.append(cnt_scan)

    # Alunos por Grupo Musical
    from alunos.models import GrupoMusical, Aluno
    grupos_data = []
    for g in GrupoMusical.objects.all():
        cnt = Aluno.objects.filter(grupo=g).count()
        if cnt > 0:
            grupos_data.append({"label": g.nome, "value": cnt})

    com_rfid = Instrumento.objects.filter(rfid__isnull=False).exclude(rfid="").count()
    sem_rfid = total_inst - com_rfid

    # Disponibilidade por naipe
    naipes_instrumentos = []
    for n in Naipe.objects.all():
        total_n = n.instrumentos.count()
        disp_n = n.instrumentos.filter(ativo=True).exclude(id__in=emprestados_ids).count()
        naipes_instrumentos.append({
            "naipe": n.nome,
            "disp": disp_n,
            "total": total_n
        })

    # Alunos com mais empréstimos
    from django.db.models import Count
    alunos_mais_emp_data = []
    ranking = (
        InstrumentoEmprestimo.objects.values('aluno__nome')
        .annotate(total_emp=Count('id'))
        .order_by('-total_emp')[:3]
    )
    for item in ranking:
        if item['aluno__nome']:
            alunos_mais_emp_data.append({
                "nome": item['aluno__nome'],
                "total": item['total_emp']
            })
        
    return JsonResponse({
        "instrumentos": {
            "total": total_inst,
            "disponiveis": disponiveis_count,
            "emprestados": emprestados_count,
            "inativos": inativos_inst,
            "pertence_assoc": Instrumento.objects.filter(pertence_associacao=True).count(),
            "com_rfid": com_rfid,
            "sem_rfid": sem_rfid,
        },
        "emprestimos": {
            "ativos": ativos_prestamos,
            "vencidos": vencidos_count,
            "devolvidos": devolvidos_prestamos,
            "total": total_prestamos
        },
        "alunos": {
            "total": total_alunos,
            "naipes": naipes_dict
        },
        "ensaios": {
            "total": total_ensaios,
            "ativos": ativos_ensaios,
            "cancelados": cancelados_ensaios
        },
        "condicoes": {
            "labels": ["Ótimo", "Bom", "Regular", "Ruim"],
            "data": [cond_otimo, cond_bom, cond_regular, cond_ruim]
        },
        "atividade": {
            "labels": labels_atividade,
            "emprestimos": emprestimos_counts,
            "scans": scans_counts
        },
        "grupos": grupos_data,
        "alertas": alertas,
        "proximo_ensaio": prox_ensaio_data,
        "instrumentos_naipes": naipes_instrumentos,
        "alunos_mais_emprestimos": alunos_mais_emp_data,
    })


@iot_api_required
def api_listar_instrumentos(request):
    instrumentos = Instrumento.objects.all()
    data = []
    
    emprestimos_ativos = {
        emp.instrumento_id: emp
        for emp in InstrumentoEmprestimo.objects.filter(devolvido=False).select_related('aluno')
    }
    
    for inst in instrumentos:
        emprestimo = emprestimos_ativos.get(inst.id)
        disponivel = emprestimo is None
        emprestado_para = f"{emprestimo.aluno.nome} {emprestimo.aluno.sobrenome}" if emprestimo else None
        
        naipe_rel = inst.naipes.first()
        naipe_nome = naipe_rel.nome if naipe_rel else "Sem Naipe"
        
        data.append({
            "id": inst.id,
            "identificador": inst.identificador,
            "nome": inst.nome,
            "naipe": naipe_nome,
            "rfid": inst.rfid,
            "disponivel": disponivel,
            "ativo": inst.ativo,
            "pertence_associacao": inst.pertence_associacao,
            "emprestado_para": emprestado_para,
            "condicao": inst.condicao
        })
        
    return JsonResponse(data, safe=False)


@iot_api_required
def api_buscar_rfid(request, rfid):
    try:
        inst = Instrumento.objects.get(rfid=rfid)
    except Instrumento.DoesNotExist:
        return JsonResponse({"erro": "Instrumento nao encontrado"}, status=404)
        
    emprestimo = InstrumentoEmprestimo.objects.filter(instrumento=inst, devolvido=False).first()
    disponivel = emprestimo is None
    emprestado_para = f"{emprestimo.aluno.nome} {emprestimo.aluno.sobrenome}" if emprestimo else None
    
    naipe_rel = inst.naipes.first()
    naipe_nome = naipe_rel.nome if naipe_rel else "Sem Naipe"
    
    return JsonResponse({
        "id": inst.id,
        "identificador": inst.identificador,
        "nome": inst.nome,
        "naipe": naipe_nome,
        "rfid": inst.rfid,
        "disponivel": disponivel,
        "ativo": inst.ativo,
        "pertence_associacao": inst.pertence_associacao,
        "emprestado_para": emprestado_para,
        "condicao": inst.condicao
    })


@csrf_exempt
@iot_api_required
def api_registrar_retirada(request):
    if request.method == "POST":
        try:
            if request.content_type == 'application/json':
                body = json.loads(request.body)
                rfid = body.get('rfid')
                aluno_id = body.get('aluno_id')
            else:
                rfid = request.POST.get('rfid')
                aluno_id = request.POST.get('aluno_id')
                
            from alunos.models import Aluno
            inst = Instrumento.objects.get(rfid=rfid)
            aluno = Aluno.objects.get(id=aluno_id)
            
            if InstrumentoEmprestimo.objects.filter(instrumento=inst, devolvido=False).exists():
                return JsonResponse({"erro": "Instrumento ja esta emprestado"}, status=400)
                
            emprestimo = InstrumentoEmprestimo.objects.create(
                instrumento=inst,
                aluno=aluno,
                data_emprestimo=timezone.now().date(),
                devolvido=False
            )
            return JsonResponse({
                "sucesso": True,
                "msg": "Retirada registrada com sucesso",
                "id": emprestimo.id
            })
        except Instrumento.DoesNotExist:
            return JsonResponse({"erro": "Instrumento com esta tag nao cadastrado"}, status=404)
        except Aluno.DoesNotExist:
            return JsonResponse({"erro": "Aluno nao encontrado"}, status=404)
        except Exception as e:
            return JsonResponse({"erro": str(e)}, status=500)
            
    return JsonResponse({"erro": "Metodo nao permitido"}, status=405)


@csrf_exempt
@iot_api_required
def api_registrar_devolucao(request):
    if request.method == "POST":
        try:
            if request.content_type == 'application/json':
                body = json.loads(request.body)
                rfid = body.get('rfid')
            else:
                rfid = request.POST.get('rfid')
                
            inst = Instrumento.objects.get(rfid=rfid)
            emprestimo = InstrumentoEmprestimo.objects.filter(instrumento=inst, devolvido=False).first()
            
            if not emprestimo:
                return JsonResponse({"erro": "Nenhum emprestimo ativo encontrado para este instrumento"}, status=404)
                
            emprestimo.devolvido = True
            emprestimo.data_devolucao = timezone.now().date()
            emprestimo.save()
            
            return JsonResponse({
                "sucesso": True,
                "msg": "Devolucao registrada com sucesso",
                "id": emprestimo.id
            })
        except Instrumento.DoesNotExist:
            return JsonResponse({"erro": "Instrumento nao cadastrado"}, status=404)
        except Exception as e:
            return JsonResponse({"erro": str(e)}, status=500)
            
    return JsonResponse({"erro": "Metodo nao permitido"}, status=405)


@iot_api_required
def api_listar_emprestimos(request):
    emprestimos = InstrumentoEmprestimo.objects.all().select_related('instrumento', 'aluno').order_by('-data_emprestimo')
    data = []
    hoje = timezone.now().date()
    
    for emp in emprestimos:
        dias_atraso = 0
        if not emp.devolvido:
            dias = (hoje - emp.data_emprestimo).days
            if dias > 14:
                dias_atraso = dias - 14
                
        data.append({
            "id": emp.id,
            "instrumento": emp.instrumento.nome,
            "identificador": emp.instrumento.identificador,
            "rfid": emp.instrumento.rfid,
            "aluno": f"{emp.aluno.nome} {emp.aluno.sobrenome}",
            "matricula": emp.aluno.matricula,
            "data_emprestimo": emp.data_emprestimo.strftime("%d/%m/%Y"),
            "data_devolucao": emp.data_devolucao.strftime("%d/%m/%Y") if emp.data_devolucao else None,
            "devolvido": emp.devolvido,
            "dias_atraso": dias_atraso
        })
        
    return JsonResponse(data, safe=False)


@csrf_exempt
@iot_api_required
def api_iot_scan(request):
    if request.method == "POST":
        try:
            if request.content_type == 'application/json':
                body = json.loads(request.body)
                rfid = body.get('rfid')
            else:
                rfid = request.POST.get('rfid')
            
            if not rfid:
                return JsonResponse({"erro": "rfid nao fornecido"}, status=400)
                
            rfid = rfid.strip()
            
            # Registrar o scan
            scan = IoTScan.objects.create(rfid=rfid)
            
            # Manter apenas os últimos 50 scans
            total_scans = IoTScan.objects.count()
            if total_scans > 50:
                old_scans = IoTScan.objects.order_by('id')[:total_scans - 50]
                for old in old_scans:
                    old.delete()
                    
            inst = Instrumento.objects.filter(rfid=rfid).first()
            vinculado = inst is not None
            
            return JsonResponse({
                "sucesso": True,
                "msg": "Leitura registrada com sucesso",
                "rfid": rfid,
                "vinculado": vinculado,
                "instrumento": inst.nome if vinculado else None
            })
        except Exception as e:
            return JsonResponse({"erro": str(e)}, status=500)
    return JsonResponse({"erro": "Metodo nao permitido"}, status=405)


@iot_api_required
def api_iot_ultimo_scan(request):
    scan = IoTScan.objects.order_by('-timestamp').first()
    if not scan:
        return JsonResponse({"rfid": None, "timestamp": None, "vinculado": False})
        
    rfid = scan.rfid
    inst = Instrumento.objects.filter(rfid=rfid).first()
    
    data = {
        "rfid": rfid,
        "timestamp": scan.timestamp.isoformat(),
        "vinculado": inst is not None
    }
    
    if inst:
        emprestimo = InstrumentoEmprestimo.objects.filter(instrumento=inst, devolvido=False).first()
        disponivel = emprestimo is None
        emprestado_para = f"{emprestimo.aluno.nome} {emprestimo.aluno.sobrenome}" if emprestimo else None
        
        naipe_rel = inst.naipes.first()
        naipe_nome = naipe_rel.nome if naipe_rel else "Sem Naipe"
        
        data["instrumento"] = {
            "id": inst.id,
            "identificador": inst.identificador,
            "nome": inst.nome,
            "naipe": naipe_nome,
            "disponivel": disponivel,
            "ativo": inst.ativo,
            "condicao": inst.condicao,
            "emprestado_para": emprestado_para,
            "pertence_associacao": inst.pertence_associacao
        }
        
    return JsonResponse(data)


@csrf_exempt
@iot_api_required
def api_iot_vincular(request):
    if request.method == "POST":
        try:
            if request.content_type == 'application/json':
                body = json.loads(request.body)
                rfid = body.get('rfid')
                instrumento_id = body.get('instrumento_id')
            else:
                rfid = request.POST.get('rfid')
                instrumento_id = request.POST.get('instrumento_id')
                
            if not rfid or not instrumento_id:
                return JsonResponse({"erro": "Campos obrigatorios ausentes"}, status=400)
                
            rfid = rfid.strip()
            
            # Verificar se a tag já está em uso
            outrom = Instrumento.objects.filter(rfid=rfid).exclude(id=instrumento_id).first()
            if outrom:
                return JsonResponse({"erro": f"Tag ja esta em uso no instrumento {outrom.nome}"}, status=400)
                
            inst = Instrumento.objects.get(id=instrumento_id)
            inst.rfid = rfid
            inst.save()
            
            return JsonResponse({
                "sucesso": True,
                "msg": f"Tag vinculada a {inst.nome} com sucesso"
            })
        except Instrumento.DoesNotExist:
            return JsonResponse({"erro": "Instrumento nao encontrado"}, status=404)
        except Exception as e:
            return JsonResponse({"erro": str(e)}, status=500)
            
    return JsonResponse({"erro": "Metodo nao permitido"}, status=405)


@csrf_exempt
@iot_api_required
def api_iot_desvincular(request):
    if request.method == "POST":
        try:
            if request.content_type == 'application/json':
                body = json.loads(request.body)
                rfid = body.get('rfid')
                instrumento_id = body.get('instrumento_id')
            else:
                rfid = request.POST.get('rfid')
                instrumento_id = request.POST.get('instrumento_id')
                
            if instrumento_id:
                inst = Instrumento.objects.get(id=instrumento_id)
            elif rfid:
                inst = Instrumento.objects.get(rfid=rfid)
            else:
                return JsonResponse({"erro": "Informe rfid ou instrumento_id"}, status=400)
                
            inst.rfid = None
            inst.save()
            return JsonResponse({
                "sucesso": True,
                "msg": f"Tag desvinculada do instrumento {inst.nome}"
            })
        except Instrumento.DoesNotExist:
            return JsonResponse({"erro": "Instrumento nao encontrado"}, status=404)
        except Exception as e:
            return JsonResponse({"erro": str(e)}, status=500)
            
    return JsonResponse({"erro": "Metodo nao permitido"}, status=405)


@iot_api_required
def api_instrumentos_sem_rfid(request):
    instrumentos = Instrumento.objects.filter(ativo=True, rfid__isnull=True) | Instrumento.objects.filter(ativo=True, rfid="")
    instrumentos = instrumentos.distinct().order_by('nome')
    
    data = []
    for inst in instrumentos:
        naipe_rel = inst.naipes.first()
        naipe_nome = naipe_rel.nome if naipe_rel else "Sem Naipe"
        data.append({
            "id": inst.id,
            "identificador": inst.identificador,
            "nome": inst.nome,
            "naipe": naipe_nome,
            "condicao": inst.condicao
        })
    return JsonResponse(data, safe=False)


@gerente_required
def iot_web_control(request):
        
    from alunos.models import Aluno
    alunos = Aluno.objects.all().order_by('nome')
    
    instrumentos_sem_rfid = Instrumento.objects.filter(ativo=True, rfid__isnull=True) | Instrumento.objects.filter(ativo=True, rfid="")
    instrumentos_sem_rfid = instrumentos_sem_rfid.distinct().order_by('nome')
    
    total_inst = Instrumento.objects.filter(ativo=True).count()
    com_rfid = Instrumento.objects.filter(ativo=True, rfid__isnull=False).exclude(rfid="").count()
    sem_rfid = total_inst - com_rfid
    
    scans_recentes = IoTScan.objects.order_by('-timestamp')[:10]
    
    context = {
        "alunos": alunos,
        "instrumentos_sem_rfid": instrumentos_sem_rfid,
        "total_inst": total_inst,
        "com_rfid": com_rfid,
        "sem_rfid": sem_rfid,
        "scans_recentes": scans_recentes,
    }
    return render(request, "instrumentos/iot_control.html", context)


import csv
from django.http import HttpResponse

@gerente_required
def exportar_instrumentos_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="instrumentos.csv"'
    
    response.write(b'\xEF\xBB\xBF')
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Nome', 'Identificador', 'Descricao', 'Pertence a Associacao', 'Ativo', 'RFID', 'Condicao'])
    
    instrumentos = Instrumento.objects.all()
    for i in instrumentos:
        pertence = 'Sim' if i.pertence_associacao else 'Não'
        ativo = 'Sim' if i.ativo else 'Não'
        writer.writerow([
            i.nome,
            i.identificador,
            i.descricao,
            pertence,
            ativo,
            i.rfid if i.rfid else 'Sem RFID',
            i.get_condicao_display()
        ])
    return response


@gerente_required
def exportar_instrumentos_excel(request):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from django.db.models import Count
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Instrumentos"
    
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="8B6E2F")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B6E2F", end_color="8B6E2F", fill_type="solid")
    data_font = Font(name=font_family, size=11)
    
    thin_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    ws['A1'] = "Inventário de Instrumentos — ABANFAR"
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 30
    
    headers = ['Nome', 'Identificador', 'Descrição', 'Associação?', 'Ativo?', 'RFID', 'Condição']
    ws.append([])
    ws.append(headers)
    
    ws.row_dimensions[3].height = 25
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    instrumentos = Instrumento.objects.all()
    for i in instrumentos:
        pertence = 'Sim' if i.pertence_associacao else 'Não'
        ativo = 'Sim' if i.ativo else 'Não'
        row_data = [
            i.nome,
            i.identificador,
            i.descricao,
            pertence,
            ativo,
            i.rfid if i.rfid else 'Sem RFID',
            i.get_condicao_display()
        ]
        ws.append(row_data)
        
    last_row = len(instrumentos) + 3
    for r in range(4, last_row + 1):
        ws.row_dimensions[r].height = 20
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.border = thin_border
            if c in [2, 4, 5, 6, 7]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    ws['I2'] = "Condição"
    ws['J2'] = "Quantidade"
    ws['I2'].font = header_font
    ws['I2'].fill = header_fill
    ws['I2'].alignment = center_align
    ws['J2'].font = header_font
    ws['J2'].fill = header_fill
    ws['J2'].alignment = center_align
    
    cond_counts = Instrumento.objects.values('condicao').annotate(count=Count('id')).order_by('-count')
    cond_mapping = dict(Instrumento._meta.get_field('condicao').choices)
    
    sum_row = 3
    for cc in cond_counts:
        cond_raw = cc['condicao']
        cond_display = cond_mapping.get(cond_raw, cond_raw)
        ws.cell(row=sum_row, column=9, value=cond_display).font = data_font
        ws.cell(row=sum_row, column=9).border = thin_border
        ws.cell(row=sum_row, column=10, value=cc['count']).font = data_font
        ws.cell(row=sum_row, column=10).alignment = center_align
        ws.cell(row=sum_row, column=10).border = thin_border
        sum_row += 1
        
    if len(cond_counts) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Instrumentos por Condição"
        chart.y_axis.title = "Quantidade"
        chart.x_axis.title = "Condição"
        
        data_ref = Reference(ws, min_col=10, min_row=2, max_row=sum_row-1)
        cats_ref = Reference(ws, min_col=9, min_row=3, max_row=sum_row-1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.legend = None
        chart.width = 16
        chart.height = 11
        ws.add_chart(chart, "L2")
        
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].column <= 7:
            max_len = 0
            for cell in col:
                if cell.row == 1 and col_letter == 'A':
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="instrumentos.xlsx"'
    return response


@gerente_required
def exportar_instrumentos_pdf(request):
    from contas.utils import render_to_pdf
    from django.utils import timezone
    
    instrumentos = Instrumento.objects.all()
    total_instrumentos = instrumentos.count()
    ativos = Instrumento.objects.filter(ativo=True).count()
    inativos = total_instrumentos - ativos
    
    context = {
        'instrumentos': instrumentos,
        'total_instrumentos': total_instrumentos,
        'ativos': ativos,
        'inativos': inativos,
        'data_geracao': timezone.now().strftime("%d/%m/%Y %H:%M"),
        'titulo': 'Relatório Geral do Inventário de Instrumentos'
    }
    
    pdf_response = render_to_pdf('instrumentos/instrumentos_pdf.html', context)
    if pdf_response:
        pdf_response['Content-Disposition'] = 'attachment; filename="instrumentos.pdf"'
        return pdf_response
    return HttpResponse("Erro ao gerar PDF", status=500)



