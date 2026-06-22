from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from contas.decorators import gerente_required, professor_required, gerente_ou_professor_required, iot_api_required
from .models import EnsaiosproModel, ApresentacaoModel, ProfessorModel
from .forms import EnsaiosForm, ApresentacaoForm, ProfessorForm, ProfessorEditForm
from django.http import HttpResponse
from alunos.models import Aluno, Naipe
from materias.models import Materia, MatriculaMateria

# home

def professores_login(request):
    return HttpResponse("login dos professores")

@login_required
def professores_home(request):
    professor = None
    total_materias = 0
    total_alunos = 0
    
    try:
        professor = request.user.perfil.professor
        total_materias = Materia.objects.filter(professor=professor).count()
        total_alunos = Aluno.objects.filter(matriculamateria__materia__professor=professor).distinct().count()
    except Exception:
        pass
        
    total_ensaios = EnsaiosproModel.objects.filter(cancelado=False).count()
    total_apresentacoes = ApresentacaoModel.objects.count()
    
    # Gráfico 1: Alunos por Naipe
    naipes_labels = []
    naipes_data = []
    for n in Naipe.objects.all():
        naipes_labels.append(n.nome)
        naipes_data.append(Aluno.objects.filter(naipe=n).count())
        
    chart_naipes = {
        "labels": naipes_labels,
        "data": naipes_data
    }
    
    # Gráfico 2: Ensaios por dia da semana
    dias_labels = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Sábado", "Domingo"]
    dias_data = []
    for dia in dias_labels:
        count = EnsaiosproModel.objects.filter(dia_semana=dia, cancelado=False).count()
        dias_data.append(count)
        
    filtered_labels = []
    filtered_data = []
    for dia, count in zip(dias_labels, dias_data):
        if count > 0:
            filtered_labels.append(dia)
            filtered_data.append(count)
            
    chart_ensaios_dias = {
        "labels": filtered_labels,
        "data": filtered_data
    }
    
    context = {
        "professor": professor,
        "total_materias": total_materias,
        "total_alunos": total_alunos,
        "total_ensaios": total_ensaios,
        "total_apresentacoes": total_apresentacoes,
        "chart_naipes": chart_naipes,
        "chart_ensaios_dias": chart_ensaios_dias,
    }
    return render(request, "professores/professores_home.html", context)


# Ensaios
# LISTAR
@login_required
def ensaios_list(request):
    ensaios = EnsaiosproModel.objects.all()
    return render(request, "professores/ensaios_list.html", {"ensaios": ensaios})

# CRIAR
@gerente_ou_professor_required
def ensaios_create(request):
    if request.method == "POST":
        form = EnsaiosForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("professores:ensaios_list")
    else:
        form = EnsaiosForm()

    return render(request, "professores/ensaios_pro.html", {"form": form})

# EDITAR
@gerente_ou_professor_required
def ensaios_edit(request, id):
    ensaio = get_object_or_404(EnsaiosproModel, id=id)
    
    if request.method == "POST":
        form = EnsaiosForm(request.POST, instance=ensaio)
        if form.is_valid():
            form.save()
            return redirect("professores:ensaios_list")
    else:
        form = EnsaiosForm(instance=ensaio)

    return render(request, "professores/ensaios_pro.html", {"form": form})

# EXCLUIR
@gerente_ou_professor_required
def ensaios_delete(request, id):
    ensaio = get_object_or_404(EnsaiosproModel, id=id)

    if request.method == "POST":
        ensaio.delete()
        return redirect("professores:ensaios_list")

    return render(request, "professores/ensaios_confirm_delete.html", {"ensaio": ensaio})

# Apresentações

@login_required
def apresentacoes_list(request):
    apresentacoes = ApresentacaoModel.objects.all()
    return render(request, "professores/apresentacoes_list.html", {"apresentacoes": apresentacoes})

# CRIAR
@gerente_ou_professor_required
def apresentacoes_create(request):
    if request.method == "POST":
        form = ApresentacaoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("professores:apresentacoes_list")
    else:
        form = ApresentacaoForm()

    return render(request, "professores/apresentacoes_form.html", {"form": form})

# EDITAR
@gerente_ou_professor_required
def apresentacoes_edit(request, id):
    apresentacao = get_object_or_404(ApresentacaoModel, id=id)

    if request.method == "POST":
        form = ApresentacaoForm(request.POST, instance=apresentacao)
        if form.is_valid():
            form.save()
            return redirect("professores:apresentacoes_list")
    else:
        form = ApresentacaoForm(instance=apresentacao)

    return render(request, "professores/apresentacoes_form.html", {"form": form})

# EXCLUIR
@gerente_ou_professor_required
def apresentacoes_delete(request, id):
    apresentacao = get_object_or_404(ApresentacaoModel, id=id)

    if request.method == "POST":
        apresentacao.delete()
        return redirect("professores:apresentacoes_list")

    return render(request, "professores/apresentacoes_confirm_delete.html", {"apresentacao": apresentacao})

#cancelar


@professor_required
def lista_ensaios_professor(request):
    ensaios = EnsaiosproModel.objects.all()
    return render(request, "professores/ensaios_professor.html", {"ensaios": ensaios})


@gerente_ou_professor_required
def cancelar_ensaio(request, id):
    ensaio = get_object_or_404(EnsaiosproModel, id=id)
    ensaio.cancelado = True
    ensaio.save()
    return redirect("professores:ensaios_professor")


@gerente_ou_professor_required
def restaurar_ensaio(request, id):
    ensaio = get_object_or_404(EnsaiosproModel, id=id)
    ensaio.cancelado = False
    ensaio.save()
    return redirect("professores:ensaios_professor")


# cadastro professor

# LISTAR
@gerente_required
@gerente_required
def professores_list(request):
    professores = ProfessorModel.objects.all()
    ativos_count = professores.filter(ativo=True).count()
    
    # Instrumentos distintos (excluindo vazios e "Não definido")
    instrumentos_count = professores.exclude(instrumento="").exclude(instrumento="Não definido").values('instrumento').distinct().count()
    
    materias_count = Materia.objects.count()
    
    contexto = {
        "professores": professores,
        "ativos_count": ativos_count,
        "instrumentos_count": instrumentos_count,
        "materias_count": materias_count,
    }
    return render(request, "professores/professores_list.html", contexto)

# CRIAR
@gerente_required
def professores_create(request):
    if request.method == "POST":
        form = ProfessorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("professores:professores_list")
    else:
        form = ProfessorForm()

    return render(request, "professores/professores_form.html", {"form": form})


# EDITAR
@gerente_required
def professores_edit(request, id):
    professor = get_object_or_404(ProfessorModel, id=id)

    if request.method == "POST":
        form = ProfessorEditForm(request.POST, instance=professor)
        if form.is_valid():
            form.save()
            return redirect("professores:professores_list")
    else:
        form = ProfessorEditForm(instance=professor)

    return render(request, "professores/professores_form.html", {"form": form})


# EXCLUIR
@gerente_required
def professores_delete(request, id):
    professor = get_object_or_404(ProfessorModel, id=id)

    if request.method == "POST":
        professor.delete()
        return redirect("professores:professores_list")

    return render(request, "professores/professores_confirm_delete.html", {"professor": professor})


@login_required
def naipes_view(request):
    naipes = Naipe.objects.all().order_by("nome")
    return render(request, "professores/naipes.html", {"naipes": naipes})


@login_required
def naipe_detalhe(request, nome):
    naipe = get_object_or_404(Naipe, nome=nome)
    alunos = Aluno.objects.filter(naipe=naipe)

    return render(request, "professores/naipe_detalhe.html", {
        "naipe": naipe,
        "alunos": alunos
    })


# ── ENDPOINTS DE API PARA O APLICATIVO ─────────────────────────────────────────

from django.http import JsonResponse

@iot_api_required
def api_listar_ensaios(request):
    ensaios = EnsaiosproModel.objects.all().order_by('data', 'inicio')
    data = []
    
    from alunos.models import Aluno
    alunos = Aluno.objects.all().select_related('grupo')
    
    # Mapear alunos por grupo musical
    alunos_por_grupo = {}
    for a in alunos:
        if a.grupo:
            g_nome = a.grupo.nome
            if g_nome not in alunos_por_grupo:
                alunos_por_grupo[g_nome] = []
            alunos_por_grupo[g_nome].append({
                "nome": f"{a.nome} {a.sobrenome}",
                "matricula": a.matricula
            })
            
    for e in ensaios:
        # Puxar alunos do grupo com o mesmo nome do ensaio
        g_alunos = alunos_por_grupo.get(e.nome, [])
        data.append({
            "id": e.id,
            "nome": e.nome,
            "dia_semana": e.dia_semana,
            "data": e.data.strftime("%Y-%m-%d"),
            "inicio": e.inicio.strftime("%H:%M"),
            "fim": e.fim.strftime("%H:%M"),
            "local": e.local,
            "cancelado": e.cancelado,
            "alunos": g_alunos
        })
        
    return JsonResponse(data, safe=False)


import csv

@gerente_required
def exportar_professores_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="professores.csv"'
    
    response.write(b'\xEF\xBB\xBF')
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Nome', 'Email', 'Telefone', 'Instrumento', 'Status'])
    
    professores = ProfessorModel.objects.all()
    for p in professores:
        status = 'Ativo' if p.ativo else 'Inativo'
        writer.writerow([
            p.nome,
            p.email,
            p.telefone,
            p.instrumento,
            status
        ])
    return response


@gerente_required
def exportar_professores_excel(request):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from django.db.models import Count
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Professores"
    
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="8B6E2F")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B6E2F", end_color="8B6E2F", fill_type="solid")
    data_font = Font(name=font_family, size=11)
    
    thin_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    ws['A1'] = "Relatório de Professores — ABANFAR"
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 30
    
    headers = ['Nome', 'E-mail', 'Telefone', 'Instrumento Principal', 'Status']
    ws.append([])
    ws.append(headers)
    
    ws.row_dimensions[3].height = 25
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    professores = ProfessorModel.objects.all()
    for p in professores:
        status = 'Ativo' if p.ativo else 'Inativo'
        row_data = [
            p.nome,
            p.email,
            p.telefone,
            p.instrumento,
            status
        ]
        ws.append(row_data)
        
    last_row = len(professores) + 3
    for r in range(4, last_row + 1):
        ws.row_dimensions[r].height = 20
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.border = thin_border
            if c in [3, 5]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    ws['G2'] = "Instrumento"
    ws['H2'] = "Quantidade"
    ws['G2'].font = header_font
    ws['G2'].fill = header_fill
    ws['G2'].alignment = center_align
    ws['H2'].font = header_font
    ws['H2'].fill = header_fill
    ws['H2'].alignment = center_align
    
    instrumento_counts = ProfessorModel.objects.values('instrumento').annotate(count=Count('id')).order_by('-count')
    sum_row = 3
    for ic in instrumento_counts:
        inst_name = ic['instrumento'] if ic['instrumento'] else 'Sem Instrumento'
        ws.cell(row=sum_row, column=7, value=inst_name).font = data_font
        ws.cell(row=sum_row, column=7).border = thin_border
        ws.cell(row=sum_row, column=8, value=ic['count']).font = data_font
        ws.cell(row=sum_row, column=8).alignment = center_align
        ws.cell(row=sum_row, column=8).border = thin_border
        sum_row += 1
        
    if len(instrumento_counts) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Professores por Instrumento"
        chart.y_axis.title = "Quantidade"
        chart.x_axis.title = "Instrumento"
        
        data_ref = Reference(ws, min_col=8, min_row=2, max_row=sum_row-1)
        cats_ref = Reference(ws, min_col=7, min_row=3, max_row=sum_row-1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.legend = None
        chart.width = 16
        chart.height = 11
        ws.add_chart(chart, "J2")
        
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].column <= 5:
            max_len = 0
            for cell in col:
                if cell.row == 1 and col_letter == 'A':
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="professores.xlsx"'
    return response


@gerente_required
def exportar_professores_pdf(request):
    from contas.utils import render_to_pdf
    from django.utils import timezone
    
    professores = ProfessorModel.objects.all()
    total_professores = professores.count()
    ativos = ProfessorModel.objects.filter(ativo=True).count()
    inativos = total_professores - ativos
    
    context = {
        'professores': professores,
        'total_professores': total_professores,
        'ativos': ativos,
        'inativos': inativos,
        'data_geracao': timezone.now().strftime("%d/%m/%Y %H:%M"),
        'titulo': 'Relatório Geral de Professores'
    }
    
    pdf_response = render_to_pdf('professores/professores_pdf.html', context)
    if pdf_response:
        pdf_response['Content-Disposition'] = 'attachment; filename="professores.pdf"'
        return pdf_response
    return HttpResponse("Erro ao gerar PDF", status=500)


